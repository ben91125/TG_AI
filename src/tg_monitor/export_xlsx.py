from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import load_settings
from .storage import SQLiteStore

LOCAL_TZ = timezone(timedelta(hours=8))
UTC = timezone.utc


@dataclass(frozen=True, slots=True)
class ExportDateRange:
    start_input: str | None
    end_input: str | None
    start_utc: datetime | None
    end_utc: datetime | None


def main() -> None:
    parser = _build_arg_parser()
    args = parser.parse_args()
    try:
        date_range = _build_date_range(args.start, args.end)
    except ValueError as exc:
        parser.error(str(exc))
    settings = load_settings()
    output_dir = Path("outputs/reports")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / _build_output_filename(date_range)

    store = SQLiteStore(settings.sqlite_path)
    try:
        workbook = Workbook()
        workbook.remove(workbook.active)

        _add_export_info_sheet(workbook, date_range)

        _add_query_sheet(
            workbook,
            "RawMessages",
            store,
            f"""
            SELECT
                rm.created_at,
                c.title AS chat_title,
                rm.chat_id,
                u.display_name AS sender_name,
                u.username AS sender_username,
                rm.sender_id,
                rm.message_id,
                rm.reply_to_message_id,
                rm.edited_at,
                rm.edit_count,
                rm.last_seen_at,
                rm.text
            FROM raw_messages rm
            LEFT JOIN chats c ON c.chat_id = rm.chat_id
            LEFT JOIN users u ON u.user_id = rm.sender_id
            {_created_at_where_clause(date_range)}
            ORDER BY rm.created_at DESC, rm.raw_message_pk DESC
            """,
            _created_at_params(date_range),
        )
        _add_query_sheet(
            workbook,
            "GameListAnalysis",
            store,
            f"""
            SELECT
                rm.created_at,
                c.title AS chat_title,
                gla.chat_id,
                gla.message_id,
                gla.is_related,
                gla.reason,
                rm.edited_at,
                rm.edit_count,
                rm.last_seen_at,
                rm.text
            FROM game_list_analysis gla
            JOIN raw_messages rm
              ON rm.chat_id = gla.chat_id
             AND rm.message_id = gla.message_id
            LEFT JOIN chats c ON c.chat_id = gla.chat_id
            {_created_at_where_clause(date_range)}
            ORDER BY rm.created_at DESC, gla.analysis_pk DESC
            """,
            _created_at_params(date_range),
        )
        _add_query_sheet(
            workbook,
            "UserReview",
            store,
            f"""
            SELECT
                rm.created_at,
                c.title AS chat_title,
                ura.user_id,
                u.display_name AS sender_name,
                u.username AS sender_username,
                ura.is_reply,
                ura.response_type,
                ura.response_type_reason,
                ura.quality_score,
                ura.quality_reason,
                rm.edited_at,
                rm.edit_count,
                rm.last_seen_at,
                rm.text
            FROM user_reply_analysis ura
            JOIN raw_messages rm
              ON rm.chat_id = ura.chat_id
             AND rm.message_id = ura.message_id
            LEFT JOIN chats c ON c.chat_id = ura.chat_id
            LEFT JOIN users u ON u.user_id = ura.user_id
            {_created_at_where_clause(date_range)}
            ORDER BY rm.created_at DESC, ura.analysis_pk DESC
            """,
            _created_at_params(date_range),
        )
        _add_query_sheet(
            workbook,
            "Chats",
            store,
            f"""
            SELECT
                c.chat_id,
                c.title,
                c.username,
                COUNT(rm.raw_message_pk) AS message_count,
                MAX(rm.created_at) AS latest_message_at
            FROM chats c
            JOIN raw_messages rm ON rm.chat_id = c.chat_id
            {_created_at_where_clause(date_range)}
            GROUP BY c.chat_id, c.title, c.username
            ORDER BY latest_message_at DESC
            """,
            _created_at_params(date_range),
        )
        _add_query_sheet(
            workbook,
            "Users",
            store,
            f"""
            SELECT
                u.user_id,
                u.display_name,
                u.username,
                COUNT(rm.raw_message_pk) AS message_count,
                MAX(rm.created_at) AS latest_message_at
            FROM users u
            JOIN raw_messages rm ON rm.sender_id = u.user_id
            {_created_at_where_clause(date_range)}
            GROUP BY u.user_id, u.display_name, u.username
            ORDER BY latest_message_at DESC
            """,
            _created_at_params(date_range),
        )

        workbook.save(output_path)
        print(output_path)
    finally:
        store.close()


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export Telegram monitor SQLite data to xlsx.")
    parser.add_argument(
        "--start",
        default="",
        help="Start time in local time. Accepted: YYYY-MM-DD, YYYY-MM-DD HH:MM, YYYY-MM-DD HH:MM:SS.",
    )
    parser.add_argument(
        "--end",
        default="",
        help="End time in local time. Accepted: YYYY-MM-DD, YYYY-MM-DD HH:MM, YYYY-MM-DD HH:MM:SS.",
    )
    return parser


def _build_date_range(start_raw: str, end_raw: str) -> ExportDateRange:
    start_input = start_raw.strip() or None
    end_input = end_raw.strip() or None
    start_utc = _parse_local_datetime(start_input, is_end=False) if start_input else None
    end_utc = _parse_local_datetime(end_input, is_end=True) if end_input else None

    if start_utc and end_utc and start_utc > end_utc:
        raise ValueError("Export start time must be earlier than or equal to end time.")

    return ExportDateRange(
        start_input=start_input,
        end_input=end_input,
        start_utc=start_utc,
        end_utc=end_utc,
    )


def _parse_local_datetime(value: str, is_end: bool) -> datetime:
    value = value.strip()
    date_formats = ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d")
    datetime_formats = (
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%dT%H:%M:%S",
    )

    for fmt in date_formats:
        try:
            parsed = datetime.strptime(value, fmt)
            if is_end:
                parsed = parsed.replace(hour=23, minute=59, second=59, microsecond=999999)
            return parsed.replace(tzinfo=LOCAL_TZ).astimezone(UTC)
        except ValueError:
            pass

    for fmt in datetime_formats:
        try:
            return datetime.strptime(value, fmt).replace(tzinfo=LOCAL_TZ).astimezone(UTC)
        except ValueError:
            pass

    try:
        parsed = datetime.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(
            "Invalid time format. Use YYYY-MM-DD, YYYY-MM-DD HH:MM, or YYYY-MM-DD HH:MM:SS."
        ) from exc

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=LOCAL_TZ)
    return parsed.astimezone(UTC)


def _build_output_filename(date_range: ExportDateRange) -> str:
    generated_at = datetime.now(LOCAL_TZ)
    if not date_range.start_input and not date_range.end_input:
        range_suffix = "_all"
    else:
        start_part = _filename_time_part(date_range.start_input) if date_range.start_input else "begin"
        end_part = _filename_time_part(date_range.end_input) if date_range.end_input else "now"
        range_suffix = f"_range_{start_part}_to_{end_part}"
    return f"tg_monitor_snapshot_{generated_at:%Y%m%d_%H%M%S}{range_suffix}.xlsx"


def _filename_time_part(value: str | None) -> str:
    if not value:
        return ""
    return (
        value.strip()
        .replace(":", "")
        .replace("/", "")
        .replace("-", "")
        .replace("T", "_")
        .replace(" ", "_")
    )


def _created_at_where_clause(date_range: ExportDateRange) -> str:
    clauses = []
    if date_range.start_utc:
        clauses.append("rm.created_at >= ?")
    if date_range.end_utc:
        clauses.append("rm.created_at <= ?")
    if not clauses:
        return ""
    return "WHERE " + " AND ".join(clauses)


def _created_at_params(date_range: ExportDateRange) -> tuple[str, ...]:
    params = []
    if date_range.start_utc:
        params.append(date_range.start_utc.isoformat())
    if date_range.end_utc:
        params.append(date_range.end_utc.isoformat())
    return tuple(params)


def _add_export_info_sheet(workbook: Workbook, date_range: ExportDateRange) -> None:
    sheet = workbook.create_sheet(title="ExportInfo")
    rows = [
        ("generated_at_local", datetime.now(LOCAL_TZ).isoformat()),
        ("timezone", "Asia/Taipei UTC+08:00"),
        ("start_input", date_range.start_input or ""),
        ("end_input", date_range.end_input or ""),
        ("start_utc", date_range.start_utc.isoformat() if date_range.start_utc else ""),
        ("end_utc", date_range.end_utc.isoformat() if date_range.end_utc else ""),
    ]
    sheet.append(["field", "value"])
    for row in rows:
        sheet.append(list(row))
    _style_sheet(sheet)


def _add_query_sheet(
    workbook: Workbook,
    title: str,
    store: SQLiteStore,
    query: str,
    params: tuple[str, ...] = (),
) -> None:
    rows = store.conn.execute(query, params).fetchall()
    sheet = workbook.create_sheet(title=title)

    headers = rows[0].keys() if rows else _empty_headers_for(title)
    sheet.append(list(headers))
    for row in rows:
        sheet.append([row[header] for header in headers])

    _style_sheet(sheet)


def _empty_headers_for(title: str) -> list[str]:
    headers_by_title = {
        "RawMessages": [
            "created_at",
            "chat_title",
            "chat_id",
            "sender_name",
            "sender_username",
            "sender_id",
            "message_id",
            "reply_to_message_id",
            "edited_at",
            "edit_count",
            "last_seen_at",
            "text",
        ],
        "GameListAnalysis": [
            "created_at",
            "chat_title",
            "chat_id",
            "message_id",
            "is_related",
            "reason",
            "edited_at",
            "edit_count",
            "last_seen_at",
            "text",
        ],
        "UserReview": [
            "created_at",
            "chat_title",
            "user_id",
            "sender_name",
            "sender_username",
            "is_reply",
            "response_type",
            "response_type_reason",
            "quality_score",
            "quality_reason",
            "edited_at",
            "edit_count",
            "last_seen_at",
            "text",
        ],
        "Chats": ["chat_id", "title", "username", "message_count", "latest_message_at"],
        "Users": ["user_id", "display_name", "username", "message_count", "latest_message_at"],
        "ExportInfo": ["field", "value"],
    }
    return headers_by_title[title]


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value or "")
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        if header == "text":
            width = 60
        elif "reason" in header:
            width = 36
        else:
            width = min(max(max_length + 2, 12), 30)
        sheet.column_dimensions[column_letter].width = width


if __name__ == "__main__":
    main()
